"""
engine/reader.py

Lector universal de archivos Excel.
"""

from pathlib import Path
from typing import List, Dict

from openpyxl import load_workbook


class ExcelReader:

    def __init__(self, file_path: str):

        self.file_path = Path(file_path)

        if not self.file_path.exists():
            raise FileNotFoundError(
                f"No existe el archivo: {self.file_path}"
            )

        self.workbook = load_workbook(
            filename=self.file_path,
            data_only=True
        )

    @property
    def sheets(self) -> List[str]:
        return self.workbook.sheetnames

    def read_sheet(self, sheet_name: str, header_row: int = 0, data_start_row: int = 1, ignore_empty_rows: bool = True) -> List[Dict]:

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(
                f"La hoja '{sheet_name}' no existe."
            )

        sheet = self.workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) <= header_row:
            return []

        headers = [
            str(h).strip() if h is not None else ""
            for h in rows[header_row]
        ]

        data = []

        for row in rows[data_start_row:]:
            if ignore_empty_rows:
                # Check if all values are None or empty strings
                is_empty = all(v is None or str(v).strip() == "" for v in row)
                if is_empty:
                    continue

            record = {}
            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else None
                record[header] = value

            data.append(record)

        return data